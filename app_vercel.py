#!/usr/bin/env python3
"""
Ragatex Product Update Web Application - Vercel Compatible Version
A modern web app for updating product status based on inventory levels
"""

from flask import Flask, render_template, request, send_file, jsonify, flash, redirect, url_for
import pandas as pd
import os
import glob
from datetime import datetime
import shutil
from werkzeug.utils import secure_filename
import tempfile
from openpyxl import load_workbook
import base64
import io

app = Flask(__name__)
app.secret_key = 'ragatex_product_update_secret_key_2024'

# Configuration for Vercel
UPLOAD_FOLDER = '/tmp/uploads'
ALLOWED_EXTENSIONS = {'csv'}
PRODUCT_STATUS_UPDATES_FOLDER = '/tmp/product_status_updates'
VLOOKUP_FILE = 'vlookup_analysis_results.csv'
TEMPLATE_FILE = 'simple-price-template(1).xlsx'

# Ensure directories exist
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(PRODUCT_STATUS_UPDATES_FOLDER, exist_ok=True)

def allowed_file(filename):
    """Check if file extension is allowed"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def find_most_recent_status_file():
    """Find the most recent file in product_status_updates folder"""
    # Look for both old and new filename patterns
    pattern1 = os.path.join(PRODUCT_STATUS_UPDATES_FOLDER, "*-product_status_update.csv")
    pattern2 = os.path.join(PRODUCT_STATUS_UPDATES_FOLDER, "*-product-status.csv")
    files = glob.glob(pattern1) + glob.glob(pattern2)
    
    if not files:
        return None
    
    # Sort by modification time (most recent first)
    most_recent = max(files, key=os.path.getmtime)
    return most_recent

def create_today_filename(source_filename):
    """Create filename with today's date and timestamp, matching the source file's naming pattern"""
    now = datetime.now()
    timestamp = now.strftime('%H%M%S')
    
    # Detect the naming pattern from the source file
    if "-product-status.csv" in source_filename:
        # New pattern: 03-09-25-product-status.csv
        return f"{now.strftime('%d-%m-%y')}-{timestamp}-product-status.csv"
    else:
        # Old pattern: 03-09-25-product_status_update.csv
        return f"{now.strftime('%d-%m-%y')}-{timestamp}-product_status_update.csv"

def process_uploaded_file(file_path):
    """Process the uploaded CSV file and return SPU IDs with inventory data"""
    try:
        df = pd.read_csv(file_path)
        
        # Check if required columns exist
        required_columns = ['SPU ID', 'price', 'inventory']
        missing_columns = [col for col in required_columns if col not in df.columns]
        
        if missing_columns:
            return None, f"Missing required columns: {', '.join(missing_columns)}"
        
        # Clean the data
        df = df.dropna(subset=['SPU ID'])
        df['SPU ID'] = df['SPU ID'].astype(str)
        df['inventory'] = pd.to_numeric(df['inventory'], errors='coerce').fillna(0)
        
        return df, None
        
    except Exception as e:
        return None, f"Error processing file: {str(e)}"

def vlookup_spu_to_product_ids(spu_ids):
    """Perform vlookup to find Product IDs for given SPU IDs"""
    try:
        # Read vlookup file
        df = pd.read_csv(VLOOKUP_FILE)
        
        # Skip first 2 rows (headers) and get actual data - data starts from row 2 (index 2)
        actual_data = df.iloc[2:].copy()
        actual_data.iloc[:, 0] = actual_data.iloc[:, 0].astype(str)  # Product ID column
        actual_data.iloc[:, 1] = actual_data.iloc[:, 1].astype(str)  # SPU ID column
        
        # Create mapping from SPU ID to Product ID
        spu_to_product = dict(zip(actual_data.iloc[:, 1], actual_data.iloc[:, 0]))
        
        # Find matches
        results = {}
        for spu_id in spu_ids:
            if spu_id in spu_to_product:
                results[spu_id] = spu_to_product[spu_id]
            else:
                results[spu_id] = None
        
        return results, None
        
    except Exception as e:
        return None, f"Error in vlookup: {str(e)}"

def update_product_status_file(source_file, target_file, spu_inventory_data, spu_to_product_mapping):
    """Update the product status file based on inventory data"""
    try:
        # Try to read with semicolon first, then comma if that fails
        delimiter = ';'
        try:
            df = pd.read_csv(source_file, sep=';', dtype=str)
        except:
            df = pd.read_csv(source_file, sep=',', dtype=str)
            delimiter = ','
        
        # Create a mapping from Product ID to inventory status
        product_to_status = {}
        for spu_id, inventory in spu_inventory_data.items():
            if spu_id in spu_to_product_mapping and spu_to_product_mapping[spu_id]:
                product_id = spu_to_product_mapping[spu_id]
                status = "Inactive" if inventory == 0 else "Active"
                product_to_status[product_id] = status
        
        # Update the status column (third column, index 2)
        status_column = df.columns[2]
        product_id_column = df.columns[0]
        
        # Update status for matching Product IDs
        updates_made = 0
        for idx, row in df.iterrows():
            if idx >= 3:  # Skip header rows (only skip first 3 rows, data starts from row 3)
                product_id = str(row[product_id_column])
                if product_id in product_to_status:
                    old_status = row[status_column]
                    new_status = product_to_status[product_id]
                    df.at[idx, status_column] = new_status
                    updates_made += 1
        
        # Clean up column headers - replace "Unnamed: 1" and "Unnamed: 2" with blank strings
        df.columns = [col if not col.startswith('Unnamed:') else '' for col in df.columns]
        
        # Save the updated file with the same delimiter
        df.to_csv(target_file, sep=delimiter, index=False, encoding='utf-8')
        
        return True, None
        
    except Exception as e:
        return False, f"Error updating file: {str(e)}"

def create_price_template_excel(spu_inventory_data, spu_to_product_mapping):
    """Create Excel file using the price template with matched Product IDs and their inventory/price data"""
    try:
        # Filter SPU IDs with inventory > 0
        spu_with_inventory = {spu_id: data for spu_id, data in spu_inventory_data.items() if data['inventory'] > 0}
        
        if not spu_with_inventory:
            return None, "No SPU IDs found with inventory > 0"
        
        # Get Product IDs for SPU IDs with inventory > 0
        matched_data = []
        for spu_id, data in spu_with_inventory.items():
            if spu_id in spu_to_product_mapping and spu_to_product_mapping[spu_id]:
                product_id = spu_to_product_mapping[spu_id]
                matched_data.append({
                    'product_id': product_id,
                    'inventory': data['inventory'],
                    'price': data['price']
                })
        
        if not matched_data:
            return None, "No Product IDs found for SPU IDs with inventory > 0"
        
        # Load the template
        wb = load_workbook(TEMPLATE_FILE)
        ws = wb.active
        
        # Clear existing data from row 4 onwards (keep headers on row 3)
        for row in range(4, ws.max_row + 1):
            for col in range(1, 4):  # Clear columns A, B, C
                ws.cell(row=row, column=col, value="")
        
        # Add matched data starting from row 4
        for idx, data in enumerate(matched_data):
            row_num = 4 + idx
            ws.cell(row=row_num, column=1, value=data['product_id'])  # Identify ID column
            ws.cell(row=row_num, column=2, value=data['inventory'])   # Inventory column
            ws.cell(row=row_num, column=3, value=data['price'])       # Price column
        
        # Create temporary file
        temp_filename = f"price_template_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        temp_path = os.path.join(UPLOAD_FOLDER, temp_filename)
        
        # Save the workbook
        wb.save(temp_path)
        
        return temp_path, None
        
    except Exception as e:
        return None, f"Error creating price template Excel: {str(e)}"

@app.route('/')
def index():
    """Main page"""
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    """Handle file upload and processing"""
    if 'file' not in request.files:
        flash('No file selected')
        return redirect(url_for('index'))
    
    file = request.files['file']
    if file.filename == '':
        flash('No file selected')
        return redirect(url_for('index'))
    
    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        file_path = os.path.join(UPLOAD_FOLDER, filename)
        file.save(file_path)
        
        try:
            # Step 1: Process uploaded file
            df, error = process_uploaded_file(file_path)
            if error:
                flash(f'Error processing file: {error}')
                return redirect(url_for('index'))
            
            # Step 2: Get SPU IDs and inventory/price data
            spu_inventory_data = {}
            for _, row in df.iterrows():
                spu_id = str(row['SPU ID'])
                spu_inventory_data[spu_id] = {
                    'inventory': row['inventory'],
                    'price': row['price']
                }
            spu_ids = list(spu_inventory_data.keys())
            
            # Step 3: Vlookup SPU IDs to Product IDs
            spu_to_product_mapping, error = vlookup_spu_to_product_ids(spu_ids)
            if error:
                flash(f'Error in vlookup: {error}')
                return redirect(url_for('index'))
            
            # Step 4: Find most recent status file
            most_recent_file = find_most_recent_status_file()
            if not most_recent_file:
                flash('No product status update files found')
                return redirect(url_for('index'))
            
            # Step 5: Create today's filename and copy file
            today_filename = create_today_filename(most_recent_file)
            today_file_path = os.path.join(PRODUCT_STATUS_UPDATES_FOLDER, today_filename)
            
            shutil.copy2(most_recent_file, today_file_path)
            
            # Step 6: Update the file with new status
            # Create inventory-only data for the existing functionality
            spu_inventory_only = {spu_id: data['inventory'] for spu_id, data in spu_inventory_data.items()}
            success, error = update_product_status_file(
                most_recent_file, 
                today_file_path, 
                spu_inventory_only, 
                spu_to_product_mapping
            )
            
            if not success:
                flash(f'Error updating file: {error}')
                return redirect(url_for('index'))
            
            # Step 7: Prepare Excel file for download (existing functionality)
            excel_filename = f"ragatex_product_update_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            excel_path = os.path.join(UPLOAD_FOLDER, excel_filename)
            
            # Convert CSV to Excel - detect delimiter first
            try:
                df_excel = pd.read_csv(today_file_path, sep=';')
            except:
                df_excel = pd.read_csv(today_file_path, sep=',')
            df_excel.to_excel(excel_path, index=False, engine='openpyxl')
            
            # Step 8: Create price template Excel file
            price_template_path, error = create_price_template_excel(spu_inventory_data, spu_to_product_mapping)
            if error:
                # Continue with just the main file if price template fails
                price_template_path = None
            
            # Clean up uploaded file
            os.remove(file_path)
            
            # Store file paths in session for download
            from flask import session
            session['excel_file'] = excel_path
            session['excel_filename'] = excel_filename
            if price_template_path:
                session['price_template_file'] = price_template_path
                session['price_template_filename'] = os.path.basename(price_template_path)
            
            # Return download page directly instead of redirecting
            return render_template('download.html')
            
        except Exception as e:
            # Clean up uploaded file if it exists
            if os.path.exists(file_path):
                os.remove(file_path)
            flash(f'Unexpected error: {str(e)}')
            return redirect(url_for('index'))
    
    else:
        flash('Invalid file type. Please upload a CSV file.')
        return redirect(url_for('index'))

@app.route('/download_files')
def download_files():
    """Download page with links to both Excel files"""
    from flask import session
    return render_template('download.html')

@app.route('/download/<file_type>')
def download_file(file_type):
    """Download specific file type"""
    from flask import session
    if file_type == 'main' and 'excel_file' in session:
        return send_file(session['excel_file'], as_attachment=True, download_name=session['excel_filename'])
    elif file_type == 'price_template' and 'price_template_file' in session:
        return send_file(session['price_template_file'], as_attachment=True, download_name=session['price_template_filename'])
    else:
        flash('File not found')
        return redirect(url_for('index'))

@app.route('/api/status')
def api_status():
    """API endpoint to check system status"""
    status = {
        'vlookup_file_exists': os.path.exists(VLOOKUP_FILE),
        'product_status_folder_exists': os.path.exists(PRODUCT_STATUS_UPDATES_FOLDER),
        'most_recent_file': find_most_recent_status_file(),
        'upload_folder_exists': os.path.exists(UPLOAD_FOLDER),
        'template_file_exists': os.path.exists(TEMPLATE_FILE)
    }
    return jsonify(status)

# For Vercel deployment
app.wsgi_app = app.wsgi_app

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5001)
